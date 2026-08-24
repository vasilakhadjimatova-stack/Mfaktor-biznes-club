"""Excel (Mbm_2026.xlsx) faylini sayt orqali yuklab import qilish.

Sozlamalar sahifasidagi «Excel'dan import» formasi shu modulni chaqiradi:
  1. «ДДС данные» varag'i → DdsRow (eski qatorlar almashtiriladi)
  2. «ДДС_2026» varag'idan hamyonlarning yil boshi qoldiqlari
  3. Kassa qayta quriladi (ddsflow.rebuild_all)
  4. To'lovlar shartnomalarga qayta taqqoslanadi (matching.run_all)

Hammasi bitta tranzaksiyada — xato bo'lsa hech narsa o'zgarmaydi.

Google Sheets havolasi berilsa (import_from_sheets) fayl serverning o'zida
yuklab olinadi — qo'lda .xlsx eksport qilish shart emas. Buning uchun jadval
«Havolaga ega bo'lganlar — ko'ruvchi» bo'lishi kerak.
"""
import io
import re
import unicodedata
import urllib.request

import openpyxl

from database import db
from models import Contract, DdsRow, Wallet

import ddsflow
import matching

SHEET_DATA = "ДДС данные"
SHEET_YEAR_PREFIX = "ДДС_"

# «ДДС_2026» qoldiq qatori nomi → hamyon kodi
OPENING_MAP = {
    "р.с мбм": "rs_mbm",
    "накт сум": "nal",
    "нал сум": "nal",
    "р.с мбм davr bank": "davr_mbm",
    "$": "usd",
    "карта 2406": "uzcard2406",
    "рс mfaktor": "rs_mfaktor",
    "pc mfaktor": "rs_mfaktor",
    "карта mfaktor": "karta_mfaktor",
}


def _norm(s):
    s = unicodedata.normalize("NFKC", str(s or "")).strip().lower()
    return " ".join(s.split())


def _num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(" ", "").replace("\xa0", "")
                     .replace(",", "."))
    except ValueError:
        return 0.0


# Havoladan faqat jadval ID'si olinadi — yuklash manzilini o'zimiz quramiz,
# shunda forma orqali boshqa saytga so'rov yuborib bo'lmaydi.
_SHEETS_ID = re.compile(r"docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]{20,})")


def import_from_sheets(link):
    """Google Sheets havolasidan .xlsx ni yuklab olib, importni bajaradi."""
    m = _SHEETS_ID.search(link or "")
    if not m:
        return {"error": "Havola noto'g'ri — u docs.google.com/spreadsheets/"
                         "d/… ko'rinishida bo'lishi kerak (jadvalning o'z "
                         "manzilini nusxalang)."}
    url = (f"https://docs.google.com/spreadsheets/d/{m.group(1)}"
           f"/export?format=xlsx")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            data = resp.read(60 * 1024 * 1024)
    except Exception as e:                              # noqa: BLE001
        return {"error": f"Google Sheets'dan yuklab bo'lmadi: {e}"}
    # Yopiq jadvalda Google xlsx o'rniga HTML kirish sahifasini qaytaradi
    if not data.startswith(b"PK"):
        return {"error": "Jadval yopiq ko'rinadi. Google Sheets'da «Ulashish» "
                         "→ «Havolaga ega bo'lganlar — ko'ruvchi» qilib "
                         "qo'ying, yoki .xlsx yuklab olib, fayl orqali "
                         "import qiling."}
    return import_workbook(io.BytesIO(data))


def import_workbook(stream):
    """Yuklangan xlsx oqimini o'qib butun zanjirni bajaradi.

    Qaytaradi: xulosa dict (sahifada ko'rsatiladi).
    """
    # read_only: katta fayl (Google butun jadvalni eksport qiladi) xotirani
    # yeb qo'ymasligi uchun — varaqlar diskdan oqim bo'lib o'qiladi.
    wb = openpyxl.load_workbook(stream, data_only=True, read_only=True)
    if SHEET_DATA not in wb.sheetnames:
        return {"error": f"Faylda «{SHEET_DATA}» varag'i topilmadi. "
                         f"Mavjud varaqlar: {', '.join(wb.sheetnames[:8])}…"}
    ws = wb[SHEET_DATA]

    # ── 1. ДДС данные → DdsRow ──
    old = DdsRow.query.count()
    # Odam qilgan qarorlar (qo'lda bog'langan / chetlatilgan) esda qoladi:
    # import ularni o'chirib yubormasligi kerak. Kalit — sana + summa + izoh.
    keep = {}
    for row in DdsRow.query.filter(DdsRow.match_status.in_(("manual", "skipped"))):
        keep[(row.ddate, round(row.amount or 0, 2),
              (row.purpose or "").strip().lower())] = (row.match_status,
                                                       row.contract_id)
    # Faqat EXCEL'dan kelgan qatorlarni almashtiramiz. Dasturda qo'lda
    # kiritilganlar (origin="app": o'tkazma tuzatishi, kofe-break va h.k.)
    # Excel'da yo'q — ularni o'chirsak, buxgalter kiritgan tuzatishlar
    # har importda yo'qolib ketardi. Shuning uchun ularга tegmaymiz.
    excel_rows = DdsRow.query.filter(DdsRow.origin != "app")
    for row in excel_rows.all():
        ddsflow.unsync_row(row)
    excel_rows.delete(synchronize_session=False)
    db.session.flush()

    added = 0
    # read_only rejimida katakka birma-bir murojaat sekin — qatorlab o'qiymiz
    for r, vals in enumerate(
            ws.iter_rows(min_row=3, max_col=8, values_only=True), start=3):
        d = vals[2]
        if d is None:
            continue
        db.session.add(DdsRow(
            rownum=r,
            ddate=d.date() if hasattr(d, "date") else d,
            amount=_num(vals[3]),
            wallet=(vals[4] or ""),
            wallet2=(vals[5] or ""),
            purpose=str(vals[6] or "").strip(),
            article=(vals[7] or ""),
        ))
        added += 1
        if added % 500 == 0:
            db.session.flush()

    # ── 2. Hamyon qoldiqlari (ДДС_2026 varag'idan) ──
    openings = 0
    year_sheet = next((n for n in wb.sheetnames
                       if n.startswith(SHEET_YEAR_PREFIX)), None)
    if year_sheet:
        ddsflow.ensure_wallets()
        ys = wb[year_sheet]
        in_bal = False
        for vals in ys.iter_rows(min_row=1, max_row=40, max_col=2,
                                 values_only=True):
            label = _norm(vals[0])
            if not label:
                continue
            if label.startswith("остаток дс на начало"):
                in_bal = True
                continue
            if in_bal:
                code = OPENING_MAP.get(label)
                if code:
                    w = Wallet.query.filter_by(code=code).first()
                    if w:
                        w.opening = _num(vals[1])
                        openings += 1
                elif label.startswith(("операционная", "поступления")):
                    break

    # ── 3-4. Kassa + moslash ──
    # Hammasi BITTA tranzaksiyada: o'chirish alohida saqlanib, keyin
    # qayta qurish uzilib qolsa, kassa bo'm-bo'sh qolib ketardi.
    reb = ddsflow.rebuild_all(commit=False)

    # odam qilgan qarorlarni qaytarib qo'yamiz
    restored = 0
    if keep:
        for row in DdsRow.query.all():
            k = (row.ddate, round(row.amount or 0, 2),
                 (row.purpose or "").strip().lower())
            st = keep.pop(k, None)
            if not st:
                continue
            status, cid = st
            if status == "skipped":
                row.match_status = "skipped"
            elif cid:
                c = db.session.get(Contract, cid)
                if c:
                    matching.apply(row, c, status="manual")
            restored += 1
        db.session.flush()

    mat = matching.run_all(commit=False)
    db.session.commit()

    return {"old": old, "added": added, "openings": openings,
            "tx": reb["made"], "tx_skipped": reb["skipped"],
            "auto": mat["auto"], "queued": mat["queued"],
            "restored": restored}
