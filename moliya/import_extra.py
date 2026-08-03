"""
Mbm_2026.xlsx qo'shimcha varaqlari importi:
  • Kurslar + real narxlar («mbm себестоимость 2026»dagi narxlar)
  • Doimiy xarajatlar → RecurringPayment («TB 2026» ro'yxati)
  • «byudjed …» varaqlari → Budget (oylik plan, statya kesimida)

Foydalanish:
  python import_extra.py /yol/Mbm_2026.xlsx
"""
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook

from app import create_app
from database import db
from models import Budget, Course, RecurringPayment

COURSES = [
    ("Sotuv menejerlari kursi (СМК)", 3_450_000),
    ("СМК Online", 3_000_000),
    ("Sotuv bo'limi rahbari (РОП)", 12_000_000),
    ("ТББ yo'nalishi", 15_000_000),
]

# TB 2026 «Jami o'zgarmas xarajatlar» ro'yxati (oylik, so'm)
RECURRING = [
    ("Xodimlar ish haqi (fiksa)", 78_400_000, 5, "Зарплата МБМ"),
    ("Dividend (fiksa)", 20_000_000, 13, "Дивиденды"),
    ("Soliqlar (oylik)", 15_000_000, 3, "Налог/дивиденд Зп"),
    ("Elektr energiya", 10_000_000, 12, "Коммунальные услуги"),
    ("Obed sotrudniki", 10_000_000, 1, "Обед сотрудников"),
    ("Internet/IP telefon", 5_862_000, 10, "Интернет/IP-телефония"),
    ("CRM", 5_500_000, 10, "CRM OnlinePBX"),
    ("Bank xizmati", 4_000_000, 25, "Комиссия банка"),
    ("Buxgalteriya arenda", 2_772_000, 12, "Аренда"),
    ("Boshqa xarajatlar", 2_000_000, 20, "Прочие расходы"),
]

# Har varaqdan FAQAT o'z oyi qatorlari olinadi — varaqlar bir-birining
# oylarini takrorlaydi (masalan, "byudjed aprel"da mart sanalari ham bor)
BUDGET_SHEETS = {"byudjed fevral": 2, "byudjed mart": 3,
                 "byudjed aprel": 4, "byudjed may": 5}


def norm(s):
    s = unicodedata.normalize("NFKC", str(s or "")).strip().lower()
    return " ".join(s.split())


def to_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except ValueError:
        return 0.0


def budget_cat(statya, comment):
    """byudjed varag'idagi erkin statya → dastur statyasi."""
    s, c = norm(statya), norm(comment)
    if "spiker" in s or "gonorar" in s:
        if "rop" in c:
            return "Зарплата РОП"
        if "tbb" in c or "тбб" in c:
            return "Зарплата ТББ"
        if "smk" in c or "смк" in c:
            return "Зарплата СМК"
        return "Премия"
    if s.startswith("zarplata") or s.startswith("зарплата"):
        return "Зарплата МБМ"
    if "target" in s:
        return "Таргет (реклама)"
    if "nalog" in s:
        return "Налог/дивиденд Зп"
    if "kofe" in s or "кофе" in s:
        return "Кофе-брейк"
    if "komunal" in s or "коммунал" in s:
        return "Коммунальные услуги"
    if "dividend" in s:
        return "Дивиденды"
    if "sotuv" in s or "zakup" in s:
        return "Закуп хоз. товаров"
    if "admin" in s:
        return "Аренда" if "arenda" in c else "Корпоративный расход"
    if "premiya" in s or "премия" in s:
        return "Премия"
    if "arenda" in s or "аренда" in s:
        return "Аренда"
    return "Прочие расходы"


def run(path):
    app = create_app()
    with app.app_context():
        wb = load_workbook(path, data_only=True)

        # 1. Kurslar — real narxlar bilan
        for name, price in COURSES:
            c = Course.query.filter_by(name=name).first()
            if not c:
                db.session.add(Course(name=name, base_price=price))
                print(f"Kurs: {name} → {price:,.0f}")
            else:
                c.base_price = price

        # 2. Doimiy xarajatlar
        if RecurringPayment.query.count() == 0:
            for name, amount, day, cat in RECURRING:
                db.session.add(RecurringPayment(name=name, amount=amount,
                                                pay_day=day, category=cat))
            print(f"Doimiy xarajatlar: {len(RECURRING)} ta "
                  f"(jami {sum(r[1] for r in RECURRING):,.0f}/oy)")

        # 3. Byudjet plan — byudjed varaqlaridan oylik jamlanadi
        plans = defaultdict(float)   # (yil, oy, statya) -> summa
        for sheet, sheet_month in BUDGET_SHEETS.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            # sarlavha qatorini topamiz
            hdr = None
            for r in range(1, 8):
                row_vals = [norm(ws.cell(row=r, column=c).value)
                            for c in range(1, 10)]
                if "sana" in row_vals and "statya" in row_vals:
                    hdr = r
                    cols = {v: i + 1 for i, v in enumerate(row_vals) if v}
                    break
            if not hdr:
                print(f"! {sheet}: sarlavha topilmadi, o'tkazildi")
                continue
            c_sana, c_st = cols["sana"], cols["statya"]
            c_sum = cols.get("summa per") or cols.get("summa")
            c_kom = cols.get("komentarie", c_st)
            n = 0
            for r in range(hdr + 1, ws.max_row + 1):
                d = ws.cell(row=r, column=c_sana).value
                if isinstance(d, datetime):
                    d = d.date()
                elif not isinstance(d, date):
                    continue
                if d.month != sheet_month:
                    continue
                amount = to_num(ws.cell(row=r, column=c_sum).value)
                if amount <= 0:
                    continue
                cat = budget_cat(ws.cell(row=r, column=c_st).value,
                                 ws.cell(row=r, column=c_kom).value)
                plans[(d.year, d.month, cat)] += amount
                n += 1
            print(f"{sheet}: {n} qator o'qildi")

        made = 0
        for (y, m, cat), total in plans.items():
            b = Budget.query.filter_by(year=y, month=m, category=cat,
                                       btype="expense").first()
            if not b:
                b = Budget(year=y, month=m, category=cat, btype="expense")
                db.session.add(b)
                made += 1
            b.planned = total
        db.session.commit()
        print(f"Byudjet: {made} ta yangi plan qatori "
              f"({len(plans)} oy-statya kombinatsiyasi)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    run(sys.argv[1])
