"""
Mfaktor ДДС (Mbm_2026.xlsx) → Mfaktor Moliya importeri.

Foydalanish:
  python import_dds.py /yol/Mbm_2026.xlsx [--wipe]

Nima qiladi:
  1. «ДДС данные» varag'idan HAR BIR tranzaksiyani o'qiydi
     (sana, summa, hamyon, statya, kirim/chiqim, faoliyat turi, izoh)
  2. «ДДС_2026» varag'idan hamyonlarning YIL BOSHI qoldiqlarini oladi
  3. Perevodlar (Техническая операция) activity="tech" bilan yoziladi —
     hamyon qoldig'iga ta'sir qiladi, statya hisobotlariga kirmaydi
  4. --wipe: avvalgi importni (comment markeri bo'yicha) o'chirib qayta yuklaydi
"""
import sys
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

from app import create_app
from database import db
from models import EXPENSE_CATS, INCOME_CATS, TRANSFER_CAT, Transaction, Wallet

IMPORT_MARK = "[dds-import]"

# «ДДС данные» hamyon nomi → (kod, ko'rsatiladigan nom)
WALLET_MAP = {
    "рс mbm": ("rs_mbm", "РС МБМ"),
    "наличные": ("nal", "Наличные (сум)"),
    "рс davr bank mbm": ("davr_mbm", "РС Davr bank МБМ"),
    "$": ("usd", "$ (valyuta)"),
    "uzcard 2406": ("uzcard2406", "UZCARD 2406"),
    "pc mfaktor": ("rs_mfaktor", "РС MFAKTOR"),
    "mfaktor karta": ("karta_mfaktor", "карта MFAKTOR"),
}
# «ДДС_2026» qoldiq qatori nomi → hamyon kodi
OPENING_MAP = {
    "р.с мбм": "rs_mbm",
    "накт сум": "nal",
    "нал сум": "nal",
    "р.с мбм davr bank": "davr_mbm",
    "$": "usd",
    "карта 2406": "uzcard2406",
    "рс mfaktor": "rs_mfaktor",
    "pc mfaktor": "rs_mfaktor",
    "карта mfaktor": "karta_mfaktor",
}

# Statya nomi (data varag'idagi) → dastur statyasi
CAT_MAP = {
    "поступление от клиента роп": "Поступление от клиента РОП",
    "поступление от клиента смк": "Поступление от клиента СМК",
    "поступление от клиента tbb": "Поступление от клиента ТББ",
    "поступление от клиента твв": "Поступление от клиента ТББ",
    "поступление от клиента тбб": "Поступление от клиента ТББ",
    "поступление б2б": "Поступление Б2Б",
    "мфактор поступления": "Мфактор поступления",
    "доход — долг": "Доход — долг",
    "доход - долг": "Доход — долг",
    "возврат поступления/клиент": "Возврат клиенту",
    "зарплата": "Зарплата МБМ",
    "зарплата смк": "Зарплата СМК",
    "зарплата роп": "Зарплата РОП",
    "зарплата тбб": "Зарплата ТББ",
    "премия": "Премия",
    "аренда": "Аренда",
    "налог дивиденд/зп": "Налог/дивиденд Зп",
    "дивиденды": "Дивиденды",
    "обед сотрудники": "Обед сотрудников",
    "комиссия банк": "Комиссия банка",
    "комунальные услуги": "Коммунальные услуги",
    "коммунальные услуги": "Коммунальные услуги",
    "ремонт": "Ремонт",
    "таргет": "Таргет (реклама)",
    "закуп/хоз товар": "Закуп хоз. товаров",
    "выпускное расходы": "Выпускные расходы",
    "кофе брейк": "Кофе-брейк",
    "срм online pbx": "CRM OnlinePBX",
    "crm online pbx": "CRM OnlinePBX",
    "такси": "Такси",
    "интернет/ip-телефония": "Интернет/IP-телефония",
    "интернет/iр-телефония": "Интернет/IP-телефония",
    "абонентские подписки": "Абонентские подписки",
    "хайрия": "Хайрия",
    "корпоративный расход": "Корпоративный расход",
    "прочие расходы": "Прочие расходы",
    "расход — долг": "Расход — долг",
    "расход - долг": "Расход — долг",
    "доход — перевод между счетами": TRANSFER_CAT,
    "расход — перевод между счетами": TRANSFER_CAT,
    "доход - перевод между счетами": TRANSFER_CAT,
    "расход - перевод между счетами": TRANSFER_CAT,
}


def norm(s):
    s = unicodedata.normalize("NFKC", str(s or "")).strip().lower()
    return " ".join(s.split())


def to_num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def ensure_wallets():
    for code, name in WALLET_MAP.values():
        if not Wallet.query.filter_by(code=code).first():
            db.session.add(Wallet(code=code, name=name, opening=0.0))
    db.session.flush()


def import_openings(wb):
    """«ДДС_2026»dan yanvar (2-ustun) qoldiqlarini hamyonlarga yozadi."""
    if "ДДС_2026" not in wb.sheetnames:
        print("! ДДС_2026 varag'i yo'q — ochilish qoldiqlari o'zgartirilmadi")
        return
    ws = wb["ДДС_2026"]
    in_bal = False
    for r in range(1, min(ws.max_row, 40) + 1):
        label = norm(ws.cell(row=r, column=1).value)
        if not label:
            continue
        if label.startswith("остаток дс на начало"):
            in_bal = True
            continue
        if in_bal:
            code = OPENING_MAP.get(label)
            if code:
                val = to_num(ws.cell(row=r, column=2).value)
                w = Wallet.query.filter_by(code=code).first()
                if w:
                    w.opening = val
                    print(f"  Qoldiq: {w.name} → {val:,.0f}")
            elif label.startswith(("операционная", "поступления")):
                break


def run(path, wipe):
    app = create_app()
    with app.app_context():
        wb = load_workbook(path, data_only=True)
        if "ДДС данные" not in wb.sheetnames:
            raise SystemExit("«ДДС данные» varag'i topilmadi. Varaqlar: "
                             + ", ".join(wb.sheetnames[:10]))
        ws = wb["ДДС данные"]

        if wipe:
            n = Transaction.query.filter(
                Transaction.comment.like(f"%{IMPORT_MARK}%")) \
                .delete(synchronize_session=False)
            print(f"Eski import o'chirildi: {n} yozuv")

        ensure_wallets()
        import_openings(wb)

        made = skipped = 0
        unknown = {}
        for r in range(3, ws.max_row + 1):
            amount = to_num(ws.cell(row=r, column=4).value)
            tdate = to_date(ws.cell(row=r, column=3).value)
            if amount <= 0 or not tdate:
                continue
            wallet_raw = norm(ws.cell(row=r, column=5).value)
            wallet = WALLET_MAP.get(wallet_raw)
            cat_raw = norm(ws.cell(row=r, column=8).value)
            cat = CAT_MAP.get(cat_raw)
            op_raw = norm(ws.cell(row=r, column=9).value)
            vid = norm(ws.cell(row=r, column=10).value)
            paykind = str(ws.cell(row=r, column=6).value or "").strip()
            purpose = str(ws.cell(row=r, column=7).value or "").strip()

            if not wallet:
                unknown.setdefault(f"hamyon: {wallet_raw}", 0)
                unknown[f"hamyon: {wallet_raw}"] += 1
                skipped += 1
                continue
            if not cat:
                unknown.setdefault(f"statya: {cat_raw}", 0)
                unknown[f"statya: {cat_raw}"] += 1
                skipped += 1
                continue

            operation = "kirim" if op_raw.startswith("поступ") else "chiqim"
            if cat == TRANSFER_CAT:
                activity = "tech"
            elif vid.startswith("финанс"):
                activity = "finance"
            else:
                activity = "operating"

            comment = f"{purpose} {IMPORT_MARK}".strip()
            if paykind:
                comment = f"[{paykind}] " + comment
            db.session.add(Transaction(
                tdate=tdate, wallet_code=wallet[0], operation=operation,
                amount=amount, category=cat, activity=activity,
                counterparty=purpose[:200], comment=comment))
            made += 1

        db.session.commit()
        print(f"\nTayyor: {made} tranzaksiya import qilindi, {skipped} o'tkazildi.")
        for k, v in unknown.items():
            print(f"  ! Notanish ({v} ta): {k}")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        raise SystemExit(__doc__)
    run(args[0], wipe="--wipe" in args)
