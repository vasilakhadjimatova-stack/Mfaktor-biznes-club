"""«ДДС данные» varag'ini 1:1 ko'chirish.

Ishlatish:
    python3 import_ddsdata.py /yo'l/Mbm_2026.xlsx
"""
import sys

import openpyxl

from app import create_app
from database import db
from models import DdsRow

SHEET = "ДДС данные"


def run(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"'{SHEET}' varag'i topilmadi. Mavjud: {wb.sheetnames}")
        return
    ws = wb[SHEET]

    app = create_app()
    with app.app_context():
        db.create_all()
        old = DdsRow.query.count()
        DdsRow.query.delete()
        db.session.commit()

        added = 0
        for r in range(3, ws.max_row + 1):
            d = ws.cell(r, 3).value          # Дата
            if d is None:
                continue
            ddate = d.date() if hasattr(d, "date") else d
            amount = ws.cell(r, 4).value or 0
            try:
                amount = float(amount)
            except (TypeError, ValueError):
                amount = 0.0
            row = DdsRow(
                rownum=r,
                ddate=ddate,
                amount=amount,
                wallet=(ws.cell(r, 5).value or ""),
                wallet2=(ws.cell(r, 6).value or ""),
                purpose=str(ws.cell(r, 7).value or ""),
                article=(ws.cell(r, 8).value or ""),
            )
            db.session.add(row)
            added += 1
            if added % 500 == 0:
                db.session.commit()
        db.session.commit()

        print(f"«{SHEET}»: {old} ta eski o'chirildi, {added} ta qator ko'chirildi")
        # nazorat: Excel formulalari bilan solishtirish
        inc = sum(x.amount for x in DdsRow.query.all() if x.flow == "Поступление")
        exp = sum(x.amount for x in DdsRow.query.all() if x.flow == "Выбытие")
        print(f"Поступление: {inc:,.0f}".replace(",", " "))
        print(f"Выбытие:     {exp:,.0f}".replace(",", " "))
        no_art = DdsRow.query.filter(DdsRow.article == "").count()
        if no_art:
            print(f"Diqqat: {no_art} ta qatorda «Статья» bo'sh")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
    else:
        run(sys.argv[1])
